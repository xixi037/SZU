import os

import math
from django.forms import model_to_dict

import openpyxl
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render

import time

from Innovation.exportword import write_to_middle
from Innovation.models import Users, Middle, Managers, ProInfo, Members, Status


def admin_base(request):
    return render(request, 'admin/admin_base.html')

def login(request):
    return render(request, 'admin/admin_login.html')

def check(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = Managers.objects.filter(username__exact=username,password__exact=password)
        if user:
            flag = 0
            response = HttpResponseRedirect('/admin/base')
            response.set_cookie('username', username, 3600)
            response.set_cookie('flag', flag, 3600)
            return response
        else:
            return HttpResponseRedirect('login')
    return HttpResponseRedirect('404.html')

def admin_welcome(request):
    print('来这里了')
    print(request.COOKIES.get('flag'))
    print(request.COOKIES.get('username'))
    if request.COOKIES.get('username', '') != '' and request.COOKIES.get('flag') == '0':
        print('进来啦')
        username = request.COOKIES.get('username')
        return render(request, 'admin/admin_welcome.html', {'username':username})
    return HttpResponseRedirect('/admin/login')

def manage_info(request):
    status = Status.objects.all()[0]
    status = status.mode
    infolist = []
    if status == 1:
        pass
        # infolist = Middle.objects.filter(status=1)
    elif status == 2:
        prolist = Middle.objects.filter(status=1)
        for i in prolist:
            id=i.pro_id
            records = ProInfo.objects.filter(id=id)
            for k in records:
                dict = model_to_dict(k)
                # infolist.append(i)
                leader_id = k.leader_id
                user = Users.objects.filter(username=leader_id)
                for j in user:
                    name = j.name
                dict['name']=name
                dict['export_time'] = i.export_time
                infolist.append(dict)
    print(infolist)
    count = len(infolist)
    print('列表长度：'+str(count))
    page = int(math.ceil(count / 20))
    pageIndex = int(request.GET.get('pageIndex', 1))
    print(pageIndex)
    info = infolist[(pageIndex - 1) * 20:20 * pageIndex]

    return render(request, 'admin/infolist.html', {'infolist':info, 'pageIndex':pageIndex, 'count':count, 'page':page})

def manage_user(request):
    alluser = Users.objects.all()
    count = alluser.count()
    page = int(math.ceil(count/20))
    print(count)
    pageIndex = int(request.GET.get('pageIndex',1))
    print(pageIndex)
    userlist = Users.objects.all()[(pageIndex-1)*20:20*pageIndex]
    return render(request, 'admin/userlist.html', {'userlist':userlist, 'pageIndex':pageIndex, 'count':count, 'page':page})

def manage_status(request):
    status = Status.objects.all()[0]
    status = status.mode
    return render(request, 'admin/status.html', {'status':status})

def manage_date(request):
    status = Status.objects.all()[0]
    date = str(status.date)
    print('date'+str(date))
    return render(request, 'admin/date.html', {'date':date})

def save_user(request):
    if 'stuID' in request.GET:
        print('----------------')
        stuID = request.GET['stuID']
        name = request.GET['name']
        sex = request.GET['sex']
        major = request.GET['major']
        institute = request.GET['institute']
        classID = request.GET['classID']
        pageIndex = request.GET.get('pageIndex','1')
        record = Users.objects.filter(username=stuID)

        if not record:
            print('不存在')
            Users.objects.create(username=stuID, password=stuID[-6:], name=name, sex=sex,major=major,institute=institute,classID=classID)
        else:
            print('已存在')
            Users.objects.filter(username=stuID).update(name=name,sex=sex,major=major,institute=institute,classID=classID)
    return HttpResponseRedirect('/admin/base/userlist?pageIndex='+str(pageIndex))

def del_name(request):
    if request.method == 'GET':
        stuID = request.GET.get('stuID')
        pageIndex = request.GET.get('pageIndex', '1')
        Users.objects.filter(username=stuID).delete()
        if ProInfo.objects.filter(leader_id=stuID):
            pro_objects = ProInfo.objects.filter(leader_id=stuID)
            for i in pro_objects:
                id = i.id
            ProInfo.objects.filter(leader_id=stuID).delete()
            Members.objects.filter(pro_id=id).delete()
            if Middle.objects.filter(leader_id=stuID):
                Middle.objects.filter(leader_id=stuID).delete()
    return HttpResponseRedirect('/admin/base/userlist?pageIndex='+str(pageIndex))

def add_name(request):
    if request.method == 'GET':
        stuID = request.GET.get('stuID')
        name = request.GET['name']
        sex = request.GET['sex']
        major = request.GET['major']
        institute = request.GET['institute']
        classID = request.GET['classID']
        pageIndex = request.GET.get('pageIndex', '1')
        record = Users.objects.filter(username=stuID)
        if not record:
            print('不存在')
            Users.objects.create(username=stuID, password=stuID[-6:], name=name, sex=sex,
                                 major=major,institute = institute,classID=classID)
        else:
            print('已存在')
            Users.objects.filter(username=stuID).update(name=name, sex=sex, major=major, institute = institute,classID=classID)

    return HttpResponseRedirect('/admin/base/userlist?pageIndex='+str(pageIndex))

def userlist_model(request):
    print('进来了')
    filename = '用户信息-模板.xlsx'
    path = os.getcwd() + os.sep + 'models'
    file = os.path.join(path, filename)
    print(file)
    return HttpResponse(file.encode('utf-8'))

def getfile(request):
    if request.FILES.get('file','') != '':
        file_obj = request.FILES.get('file')
        pageIndex = request.GET.get('pageIndex', '1')
        path = os.getcwd() + os.sep + 'models'
        savename = '上传-用户名单.xlsx'
        filepath = os.path.join(path, savename)
        print(filepath)
        dest = open(filepath,'wb+')
        dest.write(file_obj.read())
        dest.close()

        wb = openpyxl.load_workbook(filepath)
        sheet_names = wb.get_sheet_names()
        if 'Sheet1' in sheet_names:
            ws = wb.get_sheet_by_name('Sheet1')
            # 检查该条记录是否已经存在
            max_row = ws.max_row
            print(max_row)
            for index in range(2, max_row + 1):
                username = str(ws['A' + str(index)].value).strip()
                password = username[-6:]
                name = ws['B' + str(index)].value.strip()
                sex = ws['C' + str(index)].value.strip()
                major = ws['D' + str(index)].value.strip()
                institute = str(ws['E' + str(index)].value).strip()
                classID = str(ws['F' + str(index)].value).strip()
                print(username)
                print(name)
                print(sex)
                print(major)
                print(classID)
                if Users.objects.filter(username=username):
                    print('已经有了这个用户')
                    pass
                else:
                    Users.objects.create(username=username,password=password,name=name,sex=sex,major=major,institute=institute,classID=classID)

        else:
            return HttpResponse('上传文件格式有误！')
        wb.save(filepath)
        return HttpResponse('success')
    return HttpResponseRedirect('404.html')

def change_mode(request):
    if request.GET.get('status'):
        status = int(request.GET.get('status'))
        Status.objects.all().update(mode=status)
        return render(request, 'migrations/status.html', {'status': status})
    return HttpResponseRedirect('404.html')

def change_date(request):
    if request.GET.get('due_date',''):
        date = str(request.GET.get('due_date'))
        print(date)
        Status.objects.all().update(date=date)
        return render(request, 'admin/date.html', {'date': date})
    return HttpResponseRedirect('404.html')

def tolist(source):
    list = []
    for item in source:
        list.append(item.get('value'))
    return list

def export(request):
    if request.POST.get('choicelist','') != '':
        choicelist = eval(request.POST.get('choicelist'))[1:]
        print(choicelist)
        choicelist = tolist(choicelist)
        statuslist = []
        for i in choicelist:
            print('i:'+i)
            status = write_to_middle(i)
            print('status是啥：')
            print(status)
            if status=='error':
                return HttpResponse(status)
            statuslist.append(status)
        print('到底有了啥')
        # print(statuslist)
        return HttpResponse(statuslist)
        # return HttpResponse(status)
    return HttpResponse('fail')

def edit_email(request):
    return HttpResponse('还未完成 敬请期待...')

def change_password(request):
    return render(request, 'admin/admin_change_password.html')

def check_password(request):
    # print('jinlaile')
    if request.POST.get('former') != None:
        username = request.COOKIES.get('username')
        former = request.POST.get('former')
        record = Managers.objects.filter(username=username, password=former)
        if record:
            return HttpResponse('认证成功')
        else:
            return HttpResponse('密码错误')
    return HttpResponseRedirect('404')

def change_password_op(request):
    if request.POST.get('latest') != None:
        username = request.COOKIES.get('username')
        latest = request.POST.get('latest')
        Managers.objects.filter(username=username).update(password=latest)
        return HttpResponse('success')
    return HttpResponseRedirect('404')